"""
Builds downloadable Excel reports from the admin dashboard. This is the
"see it all, download it all" counterpart to the desktop client's local
per-employee export -- this one covers however many devices you select
across every branch in one file, plus the audit history and printers.
"""
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "1a237e"
LIGHT = "e8eaf6"
GREEN = "2e7d32"
RED = "c62828"

THIN = Side(style='thin', color='B0BEC5')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")


def _style_header_row(ws, headers, row=1):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal='center', vertical='center')


def _autofit(ws, n_cols, max_width=50):
    for col_idx in range(1, n_cols + 1):
        col_letter = get_column_letter(col_idx)
        longest = 10
        for cell in ws[col_letter]:
            if cell.value:
                longest = max(longest, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(longest + 3, max_width)


def _title_block(ws, title, n_cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=15, color=NAVY)
    c.alignment = Alignment(horizontal='left')

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    c = ws.cell(row=2, column=1, value=f"Generated {timezone.now():%B %d, %Y at %I:%M %p}")
    c.font = Font(size=9, italic=True, color="757575")
    return 4  # next free row


def _ram_summary(ram_devices):
    if not ram_devices:
        return ''
    return ', '.join(f"{m.get('size', '?')} ({m.get('serial', 'no serial')})" for m in ram_devices)


def _storage_summary(storage_devices):
    if not storage_devices:
        return ''
    return ', '.join(f"{d.get('size', '?')} {d.get('type', '')} ({d.get('serial', 'no serial')})"
                      for d in storage_devices)


def build_devices_workbook(devices):
    """devices: a Device queryset/list. One row per device, flattened for
    an at-a-glance fleet view -- covers every field an IT admin would
    want without having to open each device individually."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Devices"

    headers = [
        'Hostname', 'Employee', 'Designation', 'Branch', 'Device Type',
        'Status', 'Last Seen', 'Last Sync', 'IP Address', 'MAC Address', 'OS',
        'System Model', 'CPU', 'CPU Serial', 'Motherboard', 'Motherboard Serial',
        'BIOS Serial', 'RAM Total', 'RAM Modules', 'Storage', 'Monitor', 'Monitor Serial',
        'Mouse', 'Keyboard', 'UPS', 'Asset Tag', 'Vendor', 'Purchase Date',
        'Registered', 'Active',
    ]
    start_row = _title_block(ws, "Device & Asset Inventory -- All Branches", len(headers))
    _style_header_row(ws, headers, row=start_row)

    row = start_row + 1
    for device in devices:
        asset = getattr(device, 'asset', None)
        values = [
            device.hostname,
            device.employee.name if device.employee else '',
            device.employee.designation if device.employee else '',
            device.branch.name if device.branch else '',
            device.device_type,
            'Online' if device.is_online else 'Offline',
            timezone.localtime(device.last_seen).strftime('%Y-%m-%d %H:%M') if device.last_seen else '',
            timezone.localtime(device.last_sync).strftime('%Y-%m-%d %H:%M') if device.last_sync else '',
            device.ip_address, device.mac_address, device.os_info,
            asset.cpu_name if asset else '', asset.processor if asset else '',
            asset.cpu_serial if asset else '',
            f"{asset.motherboard_manufacturer} {asset.motherboard_model}".strip() if asset else '',
            asset.motherboard_serial if asset else '',
            asset.bios_serial if asset else '',
            asset.ram_total if asset else '',
            _ram_summary(asset.ram_devices) if asset else '',
            _storage_summary(asset.storage_devices) if asset else '',
            asset.monitor_name if asset else '', asset.monitor_serial if asset else '',
            asset.mouse_name if asset else '', asset.keyboard_name if asset else '',
            asset.ups_name if asset else '',
            asset.asset_tag if asset else '', asset.vendor if asset else '',
            asset.purchase_date if asset else '',
            timezone.localtime(device.registered_at).strftime('%Y-%m-%d'),
            'Yes' if device.is_active else 'No (deactivated)',
        ]
        status_col = headers.index('Status') + 1
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=val if val not in (None, '') else '')
            c.border = BORDER
            if ci == status_col:
                c.font = Font(bold=True, color=GREEN if val == 'Online' else RED)
        row += 1

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    ws.sheet_view.showGridLines = False
    _autofit(ws, len(headers))
    return wb


def build_change_log_workbook(changes):
    """changes: an AssetChangeLog queryset/list -- the audit trail export."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Asset Change History"

    headers = ['Date/Time', 'Computer', 'Employee', 'Branch', 'Field', 'Change Type', 'Previous Value', 'New Value']
    start_row = _title_block(ws, "Asset Change History (Audit Log)", len(headers))
    _style_header_row(ws, headers, row=start_row)

    row = start_row + 1
    for change in changes:
        values = [
            timezone.localtime(change.changed_at).strftime('%Y-%m-%d %H:%M'),
            change.computer_name, change.employee_name, change.branch_name,
            change.field_name, change.change_type, change.old_value, change.new_value,
        ]
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=val if val not in (None, '') else '')
            c.border = BORDER
        row += 1

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    ws.sheet_view.showGridLines = False
    _autofit(ws, len(headers))
    return wb


def build_printers_workbook(printers):
    """Matches the company's existing Printers sheet column layout exactly
    (same as the desktop client's printer export)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Printers"

    headers = ['ID', 'Brand', 'Model', 'Serial Number', 'Date Purchased', 'Warranty Expiry', 'Branch', 'Location']
    start_row = _title_block(ws, "Printer Inventory -- All Branches", len(headers))
    _style_header_row(ws, headers, row=start_row)

    row = start_row + 1
    for p in printers:
        values = [p.id, p.brand, p.model, p.serial_number, p.date_purchased,
                  p.warranty_expiry, p.branch.name if p.branch else '', p.location]
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=val if val not in (None, '') else '')
            c.border = BORDER
        row += 1

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    ws.sheet_view.showGridLines = False
    _autofit(ws, len(headers))
    return wb


def workbook_response(workbook, filename):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response
